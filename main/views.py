from django.contrib import messages
from django.db import IntegrityError, transaction
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import openpyxl 
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate , login as login_django , logout as logout_django
from django.contrib.auth.decorators import login_required
from .models import Employee , ManpowerEntry , SiteStructure
from django.db.models import Count, Sum
from datetime import datetime , date, timedelta
import json

# allowing super user 
from django.contrib.auth.decorators import login_required, user_passes_test


# @login_required(login_url='login')
# def home(request):
#     return render (request, 'home.html')
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from datetime import date
# Import your models here
# CONSTANT: USER EMAIL TO SITE MAPPING

EMAIL_SITE_MAP = {
    "admin.bmm@brighttech.net.in": "BMM",
    "admin.slr@brighttech.net.in": "SLR",
    "admin.jr@brighttech.net.in": "JAIRAJ",
    "admin.arj@brighttech.net.in": "Arjas",
    "pm.ms@brighttech.net.in": "MSSSL",
    "admin.agni@brighttech.net.in": "AGNI",
}
@login_required(login_url='login')
def FFR(request):
    """
    FFR Ledger: Pulls site structure from SiteStructure model.
    Sends only numeric IDs to the server to prevent OOM errors.
    """
    user = request.user
    user_email = (user.email or user.username).lower().strip()
    is_superuser = user.is_superuser
    today = date.today()
    yesterday = today - timedelta(days=1)
    
    # 1. Capture Date & Site Selection
    report_date = request.GET.get('report_date') or request.POST.get('report_date') or str(yesterday)
    requested_site = request.GET.get('site_selection') or request.POST.get('site_selection') or 'ALL'

    # 2. Site Authorization
    selected_site = requested_site if is_superuser else EMAIL_SITE_MAP.get(user_email, "NONE")

    # 3. Handle POST (Saving Attendance)
    if request.method == 'POST':
        present_keys = [k for k in request.POST.keys() if k.startswith('p_')]
        
        try:
            with transaction.atomic():
                for p_key in present_keys:
                    struct_id = p_key.split('_')[1]
                    # Lookup row metadata from DB
                    struct = SiteStructure.objects.get(id=struct_id)

                    # Security validation
                    if not is_superuser:
                        allowed = EMAIL_SITE_MAP.get(user_email)
                        if allowed == "AGNI":
                            if struct.site not in ["AGNI-CCM", "AGNI-IF"]: continue
                        elif struct.site != allowed:
                            continue

                    # Safe conversion to handle empty inputs
                    def s_int(v):
                        try: return int(v) if v else 0
                        except: return 0

                    ManpowerEntry.objects.update_or_create(
                        date=report_date,
                        structure=struct,
                        defaults={
                            'site': struct.site,
                            'department': struct.department,
                            'designation': struct.designation,
                            'skill_level': struct.skill_level,
                            'scope': struct.scope,
                            'present': s_int(request.POST.get(f'p_{struct_id}')),
                            'absent': s_int(request.POST.get(f'a_{struct_id}')),
                            'weekly_off': s_int(request.POST.get(f'w_{struct_id}')),
                            'overtime': s_int(request.POST.get(f'o_{struct_id}')),
                            'remarks': request.POST.get(f'rem_{struct_id}', '')
                        }
                    )
            messages.success(request, f"Records for {selected_site} on {report_date} saved successfully.")
        except Exception as e:
            messages.error(request, f"Submission error: {str(e)}")
        
        return redirect(f'/FFR/?report_date={report_date}&site_selection={selected_site}')

    # 4. Handle GET (Preparing the Table)
    # Filter structure based on user selection
    if selected_site == "ALL":
        structure_qs = SiteStructure.objects.all()
    else:
        structure_qs = SiteStructure.objects.filter(site=selected_site)

    # Fetch daily entries and map them to the structure
    daily_entries = ManpowerEntry.objects.filter(date=report_date).select_related('structure')
    entry_map = {e.structure_id: e for e in daily_entries if e.structure_id}

    # Build optimized display list
    display_list = []
    for s in structure_qs:
        e = entry_map.get(s.id)
        display_list.append({
            'id': s.id,
            'sr_no': s.sr_no,
            'site': s.site,
            'dept': s.department,
            'desig': s.designation,
            'skill': s.skill_level,
            'scope': s.scope,
            'p': e.present if e else 0,
            'a': e.absent if e else 0,
            'w': e.weekly_off if e else 0,
            'o': e.overtime if e else 0,
            'rem': e.remarks if e else '',
            'ffr': e.ff_ratio if e else 0.0,
            'abs_pct': round((e.absent / s.scope * 100), 1) if e and s.scope > 0 else 0.0
        })
    all_structures = SiteStructure.objects.all()
    all_display_list = []
    for s in all_structures:
        e = entry_map.get(s.id)
        all_display_list.append({
            'site': s.site,
            'scope': s.scope,
            'p': e.present if e else 0,
        })

    return render(request, 'ffr.html', {
        'selected_site': selected_site,
        'report_date': report_date,
        'display_list': display_list,
        'all_display_list': all_display_list,
        'is_superuser': is_superuser
    })
@login_required(login_url='login')
def export_ffr(request):

    site_filter = request.GET.get('site_selection', 'ALL')
    date_filter = request.GET.get('report_date')

    # Try to format date as DD/MM/YYYY for the header if it comes as YYYY-MM-DD
    try:
        dt = datetime.strptime(str(date_filter), '%Y-%m-%d')
        display_date = dt.strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        display_date = date_filter

    # 1. Fetch relevant Site Structure rows
    if site_filter == "ALL":
        structure_qs = SiteStructure.objects.all()
    else:
        structure_qs = SiteStructure.objects.filter(site=site_filter)

    # 2. Fetch daily entries
    daily_entries = ManpowerEntry.objects.filter(date=date_filter).select_related('structure')
    entry_map = {e.structure_id: e for e in daily_entries if e.structure_id}

    # 3. Initialize Workbook & Styles
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FFR Report"

    # Style Definitions (Matching the screenshot: Times New Roman, Black, No background)
    header_font = Font(name='Times New Roman', bold=True, size=11, color="000000")
    data_font = Font(name='Times New Roman', size=11, color="000000")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 4. Construct Multi-row Headers (A to K)
    # Row 1
    row1 = ['Site', 'Dept', 'Designation', 'Skill', 'Scope', f'DATE : {display_date}', '', '', '', '', '']
    ws.append(row1)
    # Row 2
    row2 = ['', '', '', '', '', 'P', 'A', 'Abs%', 'W/O', 'OT', 'FFR%']
    ws.append(row2)

    # Apply Merges (A-E vertically; F-K horizontally)
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.merge_cells(f'{col}1:{col}2')
    ws.merge_cells('F1:K1')

    # Apply Header Styles
    for r in [1, 2]:
        for cell in ws[r]:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    # Helper function for exact percentage formatting shown in image (e.g. 16.67% vs 100.0%)
    def format_pct(num, den):
        if den == 0:
            return "0.0%"
        val = (num / den) * 100
        # If it's a clean 1 decimal float (like 12.5% or 100.0%)
        if round(val, 1) == round(val, 2):
            return f"{val:.1f}%"
        return f"{val:.2f}%"

    # 5. Populate Data
    for s in structure_qs:
        e = entry_map.get(s.id)
        p = e.present if e else 0
        a = e.absent if e else 0
        wo = e.weekly_off if e else 0
        ot = e.overtime if e else 0
        
        abs_p = format_pct(a, s.scope)
        ffr_p = format_pct(p, s.scope)
        
        ws.append([
            s.site, s.department, s.designation, s.skill_level, s.scope,
            p, a, abs_p, wo, ot, ffr_p
        ])

    # 6. Formatting (Bug-free Auto-width)
    for col_idx in range(1, 12):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_row=1):
            cell = row[col_idx-1]
            # Safely skip MergedCells and handle values (Fixes the AttributeError)
            if type(cell).__name__ != 'MergedCell' and cell.value:
                length = len(str(cell.value))
                if length > max_length: max_length = length
        
        # Add a bit of padding to the width
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 8), 45)

    # 7. Data Cell Styling
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = thin_border
            cell.font = data_font
            # Left align Site(1), Dept(2), Designation(3); Center align the rest
            if cell.column in [1, 2, 3]:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # 8. Return Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"FFR_Report_{site_filter}_{date_filter}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response
@login_required(login_url='login')
def export_ffr_summary(request):
    """
    Date-Range Summary: Generates a horizontal pivot table showing daily P, A, W/O, OT, FFR
    for each day in the selected range, matching the required Excel layout.
    """
    site_filter = request.GET.get('site_selection', 'ALL')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')

    if not from_date_str or not to_date_str:
        messages.error(request, "Select both From and To dates for summary.")
        return redirect('FFR')

    # 1. Convert strings to date objects and generate list of all dates in range
    d1 = datetime.strptime(from_date_str, '%Y-%m-%d').date()
    d2 = datetime.strptime(to_date_str, '%Y-%m-%d').date()
    
    # Optional safeguard against massive memory crashes (e.g., max 31 days)
    num_days = (d2 - d1).days + 1
    if num_days > 31:
        messages.error(request, "Please select a date range of 31 days or less.")
        return redirect('FFR')
        
    date_list = [d1 + timedelta(days=x) for x in range(num_days)]

    # 2. Fetch Structure
    if site_filter == "ALL":
        structure_qs = SiteStructure.objects.all().order_by('sr_no')
    else:
        structure_qs = SiteStructure.objects.filter(site=site_filter).order_by('sr_no')

    # 3. Fetch all entries within range and map them: {(structure_id, date): entry}
    entries = ManpowerEntry.objects.filter(date__range=[from_date_str, to_date_str])
    entry_map = {(e.structure_id, e.date): e for e in entries}

    # 4. Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Summary {from_date_str[-2:]} to {to_date_str[-2:]}"

    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 5. Construct Headers
    row1 = ['Site', 'Dept', 'Designation', 'Skill', 'Scope']
    row2 = ['', '', '', '', '']

    for dt in date_list:
        dt_formatted = dt.strftime('%d/%m/%Y')
        # Main header for the date spanning 6 columns
        row1.extend([f"DATE : {dt_formatted}", '', '', '', '', ''])
        # Sub-headers
        row2.extend(['P', 'A', 'Abs%', 'W/O', 'OT', 'FFR%'])

    ws.append(row1)
    ws.append(row2)

    # 6. Apply Merges for Headers
    # Merge Site, Dept, Designation, Skill, Scope vertically
    for col_idx in range(1, 6):
        c_letter = get_column_letter(col_idx)
        ws.merge_cells(f'{c_letter}1:{c_letter}2')

    # Merge the DATE headers horizontally (6 columns per date)
    start_col = 6
    for dt in date_list:
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(start_col + 5)
        ws.merge_cells(f'{start_letter}1:{end_letter}1')
        start_col += 6

    # Apply Header Styling
    for r in [1, 2]:
        for cell in ws[r]:
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border_style

    # 7. Populate Data Rows
    for s in structure_qs:
        row_data = [s.site, s.department, s.designation, s.skill_level, s.scope]
        
        for dt in date_list:
            e = entry_map.get((s.id, dt))
            if e:
                scope = s.scope
                abs_pct = f"{round((e.absent / scope * 100), 1)}%" if scope > 0 else "0.0%"
                ffr_pct = f"{round((e.present / scope * 100), 1)}%" if scope > 0 else "0.0%"
                row_data.extend([e.present, e.absent, abs_pct, e.weekly_off, e.overtime, ffr_pct])
            else:
                # Fill zeros if no data uploaded for this day
                row_data.extend([0, 0, "0.0%", 0, 0, "0.0%"])
        
        ws.append(row_data)

    # 8. Adjust Widths & Borders for Data
    max_col = 5 + (len(date_list) * 6)
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    
    for col_idx in range(6, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8

    # Apply borders to all appended data rows
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border_style
            if cell.column >= 4: # Center align from Skill onwards
                cell.alignment = Alignment(horizontal='center')

    # Return as downloadable Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Summary_{site_filter}_{from_date_str}_to_{to_date_str}.xlsx"'
    wb.save(response)
    return response
@login_required(login_url='login')
def employee_list(request):
    """
    Handles displaying the employee list and calculating the status counts
    to be displayed on the dashboard.
    """
    # 1. Fetch all employees from the database.
    employees = Employee.objects.all()

    # 2. Calculate the total count and the count for each status.
    total_employees = employees.count()
    joined_count = employees.filter(status__iexact ='joined').count()
    offered_count = employees.filter(status__iexact ='offered').count()
    selected_count = employees.filter(status__iexact ='selected').count()
    pending_count = employees.filter(status__iexact ='pending').count()
    rejected_count = employees.filter(status__iexact ='rejected').count()
    left_count = employees.filter(status__iexact ='left').count()
    
    # 3. Add all calculated numbers to the context dictionary.
    context = {
        'employees': employees,
        'total_employees': total_employees,
        'joined_count': joined_count,
        'offered_count': offered_count,
        'selected_count': selected_count,
        'pending_count': pending_count,
        'rejected_count': rejected_count,
        'left_count': left_count,
    }
    
    # 4. Render the dashboard template with the context data.
    return render(request, 'dashboard.html', context)
@login_required(login_url='login')
def add_employee(request):
    """
    Handles the creation of a new employee.
    Fixed: Explicitly handles the 'status' field when it's disabled in HTML.
    """
    if request.method == 'POST':
        # 1. Extract data from POST
        name = request.POST.get('name')
        age = request.POST.get('age')
        # Note: HTML name is 'contact', mapping it to contact_number
        contact_number = request.POST.get('contact') 
        company = request.POST.get('company')
        role = request.POST.get('role')
        
        # 2. FIX FOR NULL STATUS: 
        # If the field is disabled in HTML (non-staff), request.POST.get('status') is None.
        # We must ensure it defaults to 'pending' here.
        status = request.POST.get('status')
        if not status:
            status = 'pending'
        
        # 3. Handle File Uploads
        resume = request.FILES.get('resume')
        offer_letter = request.FILES.get('offer_letter')

        # 4. Duplicate Check
        if Employee.objects.filter(contact_number=contact_number).exists():
            messages.error(request, 'This contact number already exists. Please use a different one.')
            return render(request, 'add_employee.html')

        try:
            # 5. Create instance
            # We explicitly pass the status (which is now guaranteed to be 'pending' or higher)
            Employee.objects.create(
                name=name,
                age=age,
                contact_number=contact_number,
                company=company,
                role=role,
                status=status,
                resume=resume,
                offer_letter=offer_letter
            )
            messages.success(request, 'Employee created successfully!')
            return redirect('employee_list')

        except IntegrityError as e:
            messages.error(request, f'Database Error: {str(e)}')
            return render(request, 'add_employee.html')
        except Exception as e:
            messages.error(request, f'An unexpected error occurred: {str(e)}')
            return render(request, 'add_employee.html')

    # GET request: Display blank form
    return render(request, 'add_employee.html')
@login_required(login_url='login')
def edit_employee(request, employee_id):
    # Get the specific employee object we want to edit
    employee = get_object_or_404(Employee, id=employee_id)

    # --- This block runs when the user submits the "Update Profile" form ---
    if request.method == 'POST':
        
        # 1. Update all the text-based fields from the form
        employee.name = request.POST.get('name')
        employee.age = request.POST.get('age')
        employee.contact_number = request.POST.get('contact')
        employee.company = request.POST.get('company')
        employee.role = request.POST.get('role')
        employee.status = request.POST.get('status')
        employee.remarks = request.POST.get('remarks')

        # 2. Clean up conditional fields
        # If status is not 'rejected', clear any old remarks
        if employee.status != 'rejected':
            employee.remarks = None # or '', depending on your model
        
        # If status is not 'offered', clear any old offer letter
        if employee.status != 'offered' and employee.offer_letter:
            employee.offer_letter.delete(save=False)
            employee.offer_letter = None

        # --- 3. Handle the Offer Letter File ---
        
        # Check if the "delete_offer_letter" checkbox was ticked
        if request.POST.get('delete_offer_letter'):
            if employee.offer_letter:
                employee.offer_letter.delete(save=False) # Delete file from storage
                employee.offer_letter = None # Clear the field in the database

        # Check if a *new* offer letter was uploaded
        new_offer_letter = request.FILES.get('offer_letter')
        if new_offer_letter:
            if employee.offer_letter:
                employee.offer_letter.delete(save=False) # Delete the old one first
            employee.offer_letter = new_offer_letter # Save the new one

        # --- 4. Handle the Resume File ---
        
        # Check if the "delete_resume" checkbox was ticked
        if request.POST.get('delete_resume'):
            if employee.resume:
                employee.resume.delete(save=False)
                employee.resume = None

        # Check if a *new* resume was uploaded
        new_resume = request.FILES.get('resume')
        if new_resume:
            if employee.resume:
                employee.resume.delete(save=False)
            employee.resume = new_resume

        # --- 5. Save all changes to the database ---
        employee.save()
        
        # Redirect back to the employee list page
        return redirect('employee_list') # Make sure 'employee_list' is the name of your URL

    # --- This runs if it's a GET request (just loading the page) ---
    context = {
        'employee': employee
    }
    return render(request, 'edit_employee.html', context)

@login_required(login_url='login')
def export_to_excel(request):
    """
    Handles the logic for exporting the employee list to an .xlsx file.
    Includes clickable hyperlinks for uploaded resumes.
    """
    # 1. Fetch all employee data
    employees = Employee.objects.all()
    
    # 2. Create a new Excel workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Employee List'
    
    # 3. Define the table headers
    headers = [
        'REF_ID',
        'CREATED_AT',
        'NAME', 
        'AGE', 
        'CONTACT_NUMBER', 
        'DESIGNATION', 
        'SITE', 
        'STATUS',
        'RESUME', 
        'UPDATED_AT'
    ]
    sheet.append(headers)
    
    # Style the header row (Bold)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    
    # 4. Loop through employees and add their data as rows
    for index, employee in enumerate(employees, start=2):
        resume_cell_value = "No File"
        has_resume = False
        
        if employee.resume:
            # IMPORTANT: Excel needs the full absolute URL (http://...)
            resume_url = request.build_absolute_uri(employee.resume.url)
            resume_cell_value = f'=HYPERLINK("{resume_url}", "View Resume")'
            has_resume = True
        
        row = [
            employee.ref_id,
            employee.created_at.strftime('%Y-%m-%d'),
            employee.name,
            employee.age,
            employee.contact_number,
            employee.role,
            employee.company,
            employee.status,
            resume_cell_value,
            employee.updated_at.strftime('%Y-%m-%d')
        ]
        sheet.append(row)
        
        # If a resume exists, style that specific cell to look like a hyperlink
        if has_resume:
            resume_cell = sheet.cell(row=index, column=9) # Column 9 is 'RESUME'
            resume_cell.font = Font(color="0000FF", underline="single")

    # 5. Create the HttpResponse object with the correct headers for an Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # This header tells the browser to treat the response as a file download
    response['Content-Disposition'] = 'attachment; filename="RMS_Database.xlsx"'
    
    # 6. Save the workbook to the response and return it
    workbook.save(response)
    
    return response

@login_required(login_url='login')
@user_passes_test(lambda u: u.is_superuser, login_url='employee_list')
def delete_employee(request, employee_id):
    """
    Deletes an employee from the database.
    """
    # Retrieve the specific employee object, or return a 404 error if not found.
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Delete the employee record.
    employee.delete()
    
    # Redirect to the employee list page after deletion.
    return redirect('employee_list')

def login_view(request):

    if request.method == 'POST':
        email = request.POST['email']
        # username = request.POST['username']
        password = request.POST['password']

        user = authenticate( request, username = email, password = password)

        if user is not None:
            login_django(request, user)
            # fname = user.first_name
            return redirect('employee_list')
        else:
            messages.error(request, 'The entered Email or password is incorrect.')
            return redirect('login')

    return render(request, 'authentication/login.html')


def logout(request):
    logout_django(request)
    return redirect('login')


SUMMARY_STATUSES = ['selected', 'offered', 'rejected', 'joined', 'pending', 'left']

@require_POST
@csrf_exempt 
def monthly_summary_api(request):
    """
    API endpoint to filter Employee data by created_at date range and company (site),
    and return the count of each status.
    """
    try:
        # Load JSON data from the request body
        data = json.loads(request.body)
        from_date_str = data.get('from_date')
        to_date_str = data.get('to_date')
        site = data.get('site')
        
        if not all([from_date_str, to_date_str, site]):
            return JsonResponse({'error': 'Missing date or site parameters.'}, status=400)

        # Convert date strings to datetime.date objects
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date_inclusive = datetime.strptime(to_date_str, '%Y-%m-%d').date()

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format.'}, status=400)
    except ValueError:
        return JsonResponse({'error': 'Invalid date format. Expected YYYY-MM-DD.'}, status=400)
    
    try:
        # 1. Filter the base queryset: Filter by company (site) and date range (using created_at field's date)
        filtered_employees = Employee.objects.filter(
            company=site,
            created_at__date__gte=from_date,
            created_at__date__lte=to_date_inclusive
        )

        # 2. Aggregate status counts: Group by status and count the records
        status_counts_queryset = filtered_employees.values('status').annotate(count=Count('status'))
        
        # Initialize the results dictionary with all statuses set to 0
        summary_results = {status: 0 for status in SUMMARY_STATUSES}
        total_count = 0

        # 3. Populate results and calculate total
        for item in status_counts_queryset:
            # Ensure the key is lowercase to match the JavaScript expectations
            status_key = item['status'].lower()
            if status_key in summary_results:
                summary_results[status_key] = item['count']
                total_count += item['count']

        # 4. Add the total count
        summary_results['total'] = total_count

        # 5. Return JSON response
        return JsonResponse(summary_results)

    except Exception as e:
        # Catch any database or unexpected errors
        print(f"Database error during summary generation: {e}")
        return JsonResponse({'error': f'An internal server error occurred: {e}'}, status=500)